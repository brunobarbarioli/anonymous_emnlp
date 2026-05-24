"""Annotation-engine database rows and Excel export helpers."""

from __future__ import annotations

import json
import os
import re
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from core.failure_filter import (
    failure_diagnosis_text,
    has_unresolved_execution_failure,
    unresolved_failure_records,
)
from core.item_labels import item_label_aliases


MODEL_INDEX_BY_NAME = {
    "gpt-5.5": 1,
    "claude-opus-4-7": 2,
}

REPLICATION_COLUMNS = [
    "unique_id",
    "paper_title",
    "exec_success_m1",
    "exec_success_m2",
    "exec_failure_type_m1",
    "exec_failure_type_m2",
    "exec_failure_m1_note",
    "exec_failure_m2_note",
    "correspondence_exec_m1",
    "correspondence_exec_m2",
    "identified_results_m1_n",
    "identified_results_m2_n",
    "main_result_1_m1_text",
    "main_result_1_m2_text",
    "main_result_2_m1_text",
    "main_result_2_m2_text",
    "main_result_3_m1_text",
    "main_result_3_m2_text",
    "main_result_4_m1_text",
    "main_result_4_m2_text",
    "main_result_5_m1_text",
    "main_result_5_m2_text",
    "tables_m1_n",
    "tables_m2_n",
    "t1_m1_id",
    "t1_m2_id",
    "t2_m1_id",
    "t2_m2_id",
    "t3_m1_id",
    "t3_m2_id",
    "comparison_AIRE_t1_m1_n",
    "comparison_AIRE_t2_m1_n",
    "comparison_AIRE_t1_m2_n",
    "comparison_AIRE_t2_m2_n",
    "comparison_recall_AIRE_m1",
    "comparison_recall_AIRE_m2",
    "strict_coverage_m1",
    "strict_coverage_m2",
    "strict_match_rate_m1",
    "strict_match_rate_m2",
    "relaxed_coverage_m1",
    "relaxed_coverage_m2",
    "relaxed_match_rate_m1",
    "relaxed_match_rate_m2",
    "evidence_policy_m1",
    "evidence_policy_m2",
    "dominant_evidence_tier_m1",
    "dominant_evidence_tier_m2",
    "failure_diagnosis_m1",
    "failure_diagnosis_m2",
    "match_AIRE_t1_m1_n",
    "match_AIRE_t2_m1_n",
    "match_AIRE_t1_m2_n",
    "match_AIRE_t2_m2_n",
    "incon_AIRE_m1_n",
    "incon_AIRE_m2_n",
    "rob_AIRE_m1_n",
    "rob_AIRE_m2_n",
]

ALIGNMENT_COLUMNS = [
    "unique_id",
    "paper_title",
    "model",
    "incons_AIRE_nr",
    "incon_AIRE_des",
]

ROBUSTNESS_COLUMNS = [
    "unique_id",
    "paper_title",
    "model",
    "rob_AIRE_nr",
    "rob_AIRE_des",
    "rob_AIRE_cat",
    "rob_AIRE_subcat",
]

ANNOTATION_SHEETS = [
    ("database_1_replication", "annotation_replication_papers", REPLICATION_COLUMNS),
    ("database_2_alignment", "annotation_alignment_inconsistencies", ALIGNMENT_COLUMNS),
    ("database_3_robustness", "annotation_robustness_checks", ROBUSTNESS_COLUMNS),
]

_ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-?]*[ -/]*[@-~]")
_EXCEL_ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

_CLAIM_KEYWORDS = (
    "we find",
    "we show",
    "main result",
    "primary result",
    "effect",
    "impact",
    "estimate",
    "increase",
    "decrease",
    "higher",
    "lower",
    "significant",
    "evidence",
    "suggest",
)


def _sanitize_excel_cell(value: Any) -> Any:
    """Return an openpyxl-safe cell value without mutating catalog contents."""
    if not isinstance(value, str):
        return value
    value = _ANSI_ESCAPE_RE.sub("", value)
    return _EXCEL_ILLEGAL_CHAR_RE.sub(" ", value)


def resolve_model_index(model_name: str) -> Optional[int]:
    """Return the annotation model index for a known model string."""
    normalized = (model_name or "").strip().lower()
    for name, index in MODEL_INDEX_BY_NAME.items():
        if normalized == name or normalized.startswith(name):
            return index
    return None


def paper_title_from_results(results: Dict[str, Any]) -> str:
    metadata = results.get("paper_metadata") or {}
    for key in ("title", "citation", "paper_title"):
        value = str(metadata.get(key, "") or "").strip()
        if value:
            return value
    paper_path = str(results.get("paper_path", "") or "")
    if paper_path:
        return Path(paper_path).stem.replace("_", " ").strip()
    return str(results.get("paper_id", "") or "").strip()


def _split_sentences(text: str) -> List[str]:
    normalized = re.sub(r"\s+", " ", text or "").strip()
    return [
        sentence.strip()
        for sentence in re.split(r"(?<=[.!?])\s+", normalized)
        if sentence.strip()
    ]


def _claim_score(sentence: str) -> float:
    lowered = sentence.lower()
    score = 0.0
    score += sum(8.0 for keyword in _CLAIM_KEYWORDS if keyword in lowered)
    if re.search(
        r"(?i)\btable\s+(?:\d{1,3}[a-z]?|[ivxlcdm]{1,12}[a-z]?)\b",
        sentence,
    ):
        score += 10.0
    if re.search(r"[-−]?(?:\d[\d,]*(?:\.\d+)?|\.\d+)\s*(?:%|percent|percentage point|points?)?", sentence):
        score += 6.0
    if any(token in lowered for token in ("appendix", "robustness", "placebo")):
        score -= 8.0
    word_count = len(re.findall(r"[A-Za-z][A-Za-z'\-]*", sentence))
    if 8 <= word_count <= 45:
        score += 3.0
    elif word_count > 65:
        score -= 4.0
    return score


def _dedupe_sentences(sentences: Iterable[str]) -> List[str]:
    seen: set[str] = set()
    deduped: List[str] = []
    for sentence in sentences:
        cleaned = " ".join(str(sentence or "").split())
        key = re.sub(r"[^a-z0-9]+", " ", cleaned.lower()).strip()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        deduped.append(cleaned)
    return deduped


def _selected_table_ids(results: Dict[str, Any]) -> List[str]:
    selected: List[str] = []
    for entry in results.get("headline_table_selection") or []:
        table_id = str(entry.get("item_id") or entry.get("table_id") or "").strip()
        if table_id and table_id not in selected:
            selected.append(table_id)
    if not selected:
        for item in results.get("result_item_plans") or []:
            if str(item.get("item_type", "")).lower() != "table":
                continue
            table_id = str(item.get("item_id") or "").strip()
            if table_id and table_id not in selected:
                selected.append(table_id)
            if len(selected) >= 2:
                break
    if not selected:
        counts = Counter(
            str(record.get("table_name") or "").strip()
            for record in results.get("comparisons") or []
            if str(record.get("table_name") or "").strip()
        )
        selected = [table_id for table_id, _ in counts.most_common(2)]
    return selected[:2]


def _table_aliases(table_id: str) -> set[str]:
    normalized = re.sub(r"[^a-z0-9]+", "", (table_id or "").lower())
    aliases = {normalized, *item_label_aliases(table_id, table_id)}
    match = re.search(r"(?:table|tbl)(\d+)", normalized) or re.search(r"(\d+)", normalized)
    if match:
        number = match.group(1)
        aliases.update({f"table{number}", f"tbl{number}", number})
    return {alias for alias in aliases if alias}


def _same_table(left: str, right: str) -> bool:
    left_aliases = _table_aliases(left)
    right_aliases = _table_aliases(right)
    return bool(left_aliases and right_aliases and left_aliases.intersection(right_aliases))


def _existing_important_claims(
    results: Dict[str, Any],
    limit: int,
    *,
    require_full_limit: bool = True,
) -> List[Dict[str, Any]]:
    raw_claims = results.get("important_claims") or results.get("main_results") or []
    if not isinstance(raw_claims, list):
        return []
    selected_tables = _selected_table_ids(results)
    claims: List[Dict[str, Any]] = []
    for index, raw_claim in enumerate(raw_claims[:limit], start=1):
        if isinstance(raw_claim, dict):
            claim_text = str(raw_claim.get("claim_text") or raw_claim.get("text") or "").strip()
            mapped_tables = [
                str(table_id).strip()
                for table_id in (
                    raw_claim.get("mapped_tables")
                    or raw_claim.get("selected_tables")
                    or raw_claim.get("tables")
                    or []
                )
                if str(table_id).strip()
            ]
            try:
                rank = int(raw_claim.get("claim_rank") or raw_claim.get("rank") or index)
            except (TypeError, ValueError):
                rank = index
            source = str(raw_claim.get("source") or "run_payload").strip()
        else:
            claim_text = str(raw_claim or "").strip()
            mapped_tables = []
            rank = index
            source = "run_payload"
        if not claim_text:
            continue
        claims.append(
            {
                "claim_rank": rank,
                "claim_text": claim_text,
                "mapped_tables": (mapped_tables or selected_tables)[:2],
                "source": source,
            }
        )
    if require_full_limit and len(claims) < limit:
        return []
    return claims[:limit]


def build_important_claims(results: Dict[str, Any], limit: int = 5) -> List[Dict[str, Any]]:
    """Extract exactly five important claim records and map each to up to two tables."""
    model_required = str(results.get("important_claims_source") or "").lower() == "model" or (
        results.get("claims_model_generated") is True
    )
    claim_payload = results.get("claim_agent_payload") or {}
    if model_required and isinstance(claim_payload, dict):
        payload_claims = _existing_important_claims(
            {
                **results,
                "important_claims": claim_payload.get("important_claims")
                or claim_payload.get("main_results")
                or [],
            },
            limit,
            require_full_limit=False,
        )
        if payload_claims:
            return payload_claims[:limit]
    existing_claims = _existing_important_claims(
        results,
        limit,
        require_full_limit=not model_required,
    )
    if existing_claims or model_required:
        return existing_claims

    focus = results.get("headline_focus_text") or {}
    text_parts = [
        str(focus.get("abstract", "") or ""),
        str(focus.get("introduction", "") or ""),
        str((results.get("paper_metadata") or {}).get("paper_summary", "") or ""),
    ]
    sentences = _dedupe_sentences(_split_sentences(" ".join(text_parts)))
    scored = sorted(
        ((sentence, _claim_score(sentence)) for sentence in sentences),
        key=lambda item: (-item[1], sentences.index(item[0])),
    )
    selected_sentences = [sentence for sentence, score in scored if score > 0][:limit]
    if len(selected_sentences) < limit:
        for sentence in sentences:
            if sentence not in selected_sentences:
                selected_sentences.append(sentence)
            if len(selected_sentences) >= limit:
                break

    selected_tables = _selected_table_ids(results)
    table_titles = {
        str(entry.get("item_id") or ""): str(entry.get("title") or entry.get("item_id") or "")
        for entry in results.get("headline_table_selection") or []
    }
    while len(selected_sentences) < limit:
        rank = len(selected_sentences) + 1
        table_hint = selected_tables[(rank - 1) % len(selected_tables)] if selected_tables else "the selected headline table"
        selected_sentences.append(f"Main empirical claim {rank} linked to {table_titles.get(table_hint, table_hint)}.")

    claims: List[Dict[str, Any]] = []
    for rank, sentence in enumerate(selected_sentences[:limit], start=1):
        lowered = sentence.lower()
        explicit_tables = [
            table_id for table_id in selected_tables if any(alias in lowered for alias in _table_aliases(table_id))
        ]
        mapped_tables = explicit_tables or selected_tables
        claims.append(
            {
                "claim_rank": rank,
                "claim_text": sentence,
                "mapped_tables": mapped_tables[:2],
                "source": "abstract_introduction",
            }
        )
    return claims


def _failure_codes_and_note(results: Dict[str, Any]) -> tuple[str, str]:
    failure_records = unresolved_failure_records(
        results,
        prefer_existing=not bool(results.get("failure_records")),
    )
    diagnosis_note = failure_diagnosis_text(results, max_records=6, max_chars=6000)
    codes: List[str] = []
    notes: List[str] = []
    for record in failure_records[:6]:
        text = " ".join(
            str(record.get(key, "") or "")
            for key in ("severity", "stage", "tool", "likely_cause", "stderr_excerpt", "recommended_fix")
        ).lower()
        if any(token in text for token in ("unsupported_by_package", "blocked_unbound", "coverage_gap", "no package-bound")):
            codes.append("8")
            continue
        if any(token in text for token in ("missing data", "file not found", "no such file", "data")):
            codes.append("1")
        if any(token in text for token in ("package", "module", "library", "dependency")):
            codes.append("2" if "missing" in text or "unavailable" in text else "6")
        if any(token in text for token in ("syntax", "parse", "invalid syntax")):
            codes.append("3")
        if any(token in text for token in ("path", "working directory", "file order", "not found")):
            codes.append("4")
        if any(token in text for token in ("version", "deprecated", "deprecation")):
            codes.append("5")
        notes.append(str(record.get("likely_cause") or record.get("stderr_excerpt") or "").strip())
    unsupported_items = [
        item
        for item in results.get("unsupported_items") or []
        if item.get("unsupported_reason") or str(item.get("evidence_status", "")).startswith("blocked")
    ]
    if unsupported_items:
        codes.append("8")
        item_labels = ", ".join(
            str(item.get("title") or item.get("item_id") or "unknown")
            for item in unsupported_items[:6]
        )
        notes.append(
            "unsupported_by_package: selected item(s) have no package-bound step "
            f"or verified current-run artifact ({item_labels})"
        )
    if not codes and has_unresolved_execution_failure(results):
        codes.append("8")
    if not codes:
        return "", ""
    deduped_codes = sorted(set(codes), key=lambda value: int(value))
    note = diagnosis_note or " | ".join(note for note in notes if note)
    note = note[:6000]
    return "+".join(deduped_codes), note


def _comparison_counts_by_table(results: Dict[str, Any], table_id: str) -> tuple[int, int]:
    compared = 0
    matches = 0
    for record in results.get("comparisons") or []:
        record_table = str(record.get("table_name") or "")
        if not _same_table(table_id, record_table):
            continue
        compared += 1
        if record.get("match"):
            matches += 1
    return compared, matches


def _dominant_evidence_tier(results: Dict[str, Any]) -> str:
    tiers: Counter[str] = Counter()
    for record in results.get("comparisons") or []:
        if not isinstance(record, dict):
            continue
        tier = str(record.get("evidence_tier") or "").strip()
        if tier:
            tiers[tier] += 1
    if not tiers:
        return ""
    return tiers.most_common(1)[0][0]


def build_replication_update(
    results: Dict[str, Any],
    *,
    model_index: int,
    alignment_count: int,
    robustness_count: int,
) -> Dict[str, Any]:
    suffix = f"m{model_index}"
    claims = build_important_claims(results)
    table_ids = _selected_table_ids(results)
    failure_codes, failure_note = _failure_codes_and_note(results)
    compared_total = int(results.get("paper_visible_compared_total") or results.get("compared_total") or 0)
    manifest_total = int(results.get("paper_visible_manifest_total") or results.get("manifest_total") or 0)
    script_steps_completed = int(results.get("script_steps_completed") or 0)
    partial_results = bool(results.get("partial_results_available") or compared_total > 0)
    exec_success = 1 if (results.get("status") == "completed" or script_steps_completed > 0 or partial_results) else 0
    if manifest_total and compared_total >= manifest_total:
        correspondence_exec = 1
    elif compared_total > 0:
        correspondence_exec = 2
    else:
        correspondence_exec = 0

    update: Dict[str, Any] = {
        f"exec_success_{suffix}": exec_success,
        f"exec_failure_type_{suffix}": failure_codes,
        f"exec_failure_{suffix}_note": failure_note,
        f"correspondence_exec_{suffix}": correspondence_exec,
        f"identified_results_{suffix}_n": len(claims),
        f"tables_{suffix}_n": len(table_ids),
        f"t1_{suffix}_id": table_ids[0] if len(table_ids) >= 1 else "",
        f"t2_{suffix}_id": table_ids[1] if len(table_ids) >= 2 else "",
        f"t3_{suffix}_id": "",
        f"comparison_recall_AIRE_{suffix}": float(results.get("coverage_pct") or 0.0),
        f"strict_coverage_{suffix}": float(
            results.get("strict_coverage_pct")
            if results.get("strict_coverage_pct") is not None
            else results.get("coverage_pct")
            or 0.0
        ),
        f"strict_match_rate_{suffix}": float(results.get("strict_match_rate_pct") or 0.0),
        f"relaxed_coverage_{suffix}": float(
            results.get("relaxed_coverage_pct")
            if results.get("relaxed_coverage_pct") is not None
            else results.get("coverage_pct")
            or 0.0
        ),
        f"relaxed_match_rate_{suffix}": float(results.get("relaxed_match_rate_pct") or 0.0),
        f"evidence_policy_{suffix}": str(results.get("evidence_policy") or "strict_bound"),
        f"dominant_evidence_tier_{suffix}": _dominant_evidence_tier(results),
        f"failure_diagnosis_{suffix}": failure_diagnosis_text(results, max_records=3, max_chars=1800),
        f"incon_AIRE_{suffix}_n": alignment_count,
        f"rob_AIRE_{suffix}_n": robustness_count,
    }
    for table_rank in (1, 2):
        table_id = table_ids[table_rank - 1] if len(table_ids) >= table_rank else ""
        compared, matches = _comparison_counts_by_table(results, table_id) if table_id else (0, 0)
        update[f"comparison_AIRE_t{table_rank}_{suffix}_n"] = compared
        update[f"match_AIRE_t{table_rank}_{suffix}_n"] = matches
    claims_by_rank = {
        int(claim.get("claim_rank") or rank): str(claim.get("claim_text") or "").strip()
        for rank, claim in enumerate(claims, start=1)
    }
    for rank in range(1, 6):
        update[f"main_result_{rank}_{suffix}_text"] = claims_by_rank.get(rank, "")
    return update


def build_alignment_rows(
    results: Dict[str, Any],
    alignment_payload: Dict[str, Any],
    *,
    model_index: int,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    paper_id = str(results.get("paper_id") or "")
    title = paper_title_from_results(results)
    findings = alignment_payload.get("findings") or []
    for finding in findings:
        if not isinstance(finding, dict):
            continue
        status = str(finding.get("status") or "").lower()
        severity = str(finding.get("severity") or "").lower()
        message = str(
            finding.get("message")
            or finding.get("description")
            or finding.get("incon_AIRE_des")
            or ""
        ).strip()
        if not message:
            parts = [
                ("Issue", finding.get("issue")),
                ("Manuscript location", finding.get("manuscript_location")),
                ("Code location", finding.get("code_location")),
                ("What differs", finding.get("what_differs")),
                ("Why it matters", finding.get("why_it_matters")),
                ("Affected outputs", finding.get("affected_outputs")),
                ("Confidence", finding.get("confidence")),
            ]
            message = " | ".join(
                f"{label}: {value}"
                for label, value in parts
                if str(value or "").strip()
            )
        if not message or (status == "aligned" and severity in {"", "info"}):
            continue
        rows.append(
            {
                "unique_id": paper_id,
                "paper_title": title,
                "model": model_index,
                "incons_AIRE_nr": len(rows) + 1,
                "incon_AIRE_des": message,
                "severity": severity,
                "status": status,
            }
        )
    return rows


def build_robustness_rows(
    results: Dict[str, Any],
    robustness_payload: Dict[str, Any],
    *,
    model_index: int,
) -> List[Dict[str, Any]]:
    paper_id = str(results.get("paper_id") or "")
    title = paper_title_from_results(results)
    checks = list(robustness_payload.get("checks") or [])[:4]
    rows: List[Dict[str, Any]] = []
    if not checks:
        diagnosis = failure_diagnosis_text(results, max_records=3, max_chars=1800)
        if has_unresolved_execution_failure(results):
            description = (
                "Blocked: robustness checks were not proposed because the replication "
                "run has no verified current-run evidence for the selected main results."
            )
            subcategory = "replication_failure_no_verified_evidence"
        else:
            description = (
                "Blocked: robustness checks were not proposed because no structured "
                "robustness-agent output was available for this run."
            )
            subcategory = "robustness_agent_output_missing"
        if diagnosis:
            description = f"{description} Diagnosis: {diagnosis}"
        for index in range(1, 5):
            rows.append(
                {
                    "unique_id": paper_id,
                    "paper_title": title,
                    "model": model_index,
                    "rob_AIRE_nr": index,
                    "rob_AIRE_des": description,
                    "rob_AIRE_cat": "blocked",
                    "rob_AIRE_subcat": subcategory,
                    "status": "blocked",
                }
            )
        return rows
    for index, check in enumerate(checks[:4], start=1):
        if not isinstance(check, dict):
            continue
        description = str(check.get("summary") or check.get("description") or check.get("name") or "").strip()
        if not description:
            continue
        if check.get("justification"):
            description = f"{description} {check.get('justification')}".strip()
        rows.append(
            {
                "unique_id": paper_id,
                "paper_title": title,
                "model": model_index,
                "rob_AIRE_nr": index,
                "rob_AIRE_des": description,
                "rob_AIRE_cat": str(check.get("category") or "specification").strip(),
                "rob_AIRE_subcat": str(check.get("subcategory") or check.get("name") or "main_results").strip(),
                "status": str(check.get("status") or "proposed"),
            }
        )
    return rows


def export_annotation_workbook(catalog_path: str, output_path: str) -> str:
    """Export annotation tables to one XLSX workbook with three sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    with sqlite3.connect(catalog_path) as connection:
        connection.row_factory = sqlite3.Row
        for sheet_name, table_name, columns in ANNOTATION_SHEETS:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(columns)
            columns_sql = ", ".join(columns)
            if table_name == "annotation_replication_papers":
                query = f"SELECT {columns_sql} FROM {table_name} ORDER BY unique_id"
            elif table_name == "annotation_alignment_inconsistencies":
                query = (
                    f"SELECT {columns_sql} FROM {table_name} "
                    "ORDER BY unique_id, model, incons_AIRE_nr"
                )
            else:
                query = (
                    f"SELECT {columns_sql} FROM {table_name} "
                    "ORDER BY unique_id, model, rob_AIRE_nr"
                )
            rows = connection.execute(query).fetchall()
            for row in rows:
                sheet.append([_sanitize_excel_cell(row[column]) for column in columns])
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column_cells in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in column_cells)
                width = min(max(max_len + 2, 10), 55)
                sheet.column_dimensions[column_cells[0].column_letter].width = width
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(output_path)
    return os.path.abspath(output_path)


def annotation_table_counts(catalog_path: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    with sqlite3.connect(catalog_path) as connection:
        for _, table_name, _ in ANNOTATION_SHEETS:
            counts[table_name] = int(
                connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            )
    return counts


def report_presence_by_run(catalog_path: str) -> List[Dict[str, Any]]:
    with sqlite3.connect(catalog_path) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT
                r.run_id,
                r.paper_id,
                r.model_name,
                r.status,
                MAX(CASE WHEN a.role IN ('latex-report', 'replication-report') THEN 1 ELSE 0 END) AS replication_report,
                MAX(CASE WHEN a.role = 'alignment-report' THEN 1 ELSE 0 END) AS alignment_report,
                MAX(CASE WHEN a.role = 'robustness-report' THEN 1 ELSE 0 END) AS robustness_report
            FROM runs r
            LEFT JOIN artifacts a ON a.run_id = r.run_id
            WHERE r.status IN ('completed', 'partial', 'incomplete', 'blocked')
            GROUP BY r.run_id, r.paper_id, r.model_name, r.status
            ORDER BY r.paper_id, r.model_name, r.run_id
            """
        ).fetchall()
    return [dict(row) for row in rows]


def loads_json(value: Any, default: Any) -> Any:
    if value in (None, ""):
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError:
        return default
