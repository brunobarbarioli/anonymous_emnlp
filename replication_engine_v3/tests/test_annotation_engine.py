import os
import sqlite3

from openpyxl import load_workbook

from core.annotation_engine import (
    ALIGNMENT_COLUMNS,
    REPLICATION_COLUMNS,
    ROBUSTNESS_COLUMNS,
    build_important_claims,
    build_robustness_rows,
)
from core.run_context import StorageConfig
from core.storage import RunCatalog


def _paper_path(tmp_path, paper_id="10001"):
    paper_dir = tmp_path / paper_id
    paper_dir.mkdir()
    path = paper_dir / "paper.pdf"
    path.write_text("placeholder", encoding="utf-8")
    return str(path)


def _results(paper_path, status="completed"):
    return {
        "run_id": "run-test",
        "paper_id": "10001",
        "paper_path": paper_path,
        "status": status,
        "paper_metadata": {"title": "Test paper"},
        "headline_focus_text": {
            "abstract": (
                "We find that the policy increases enrollment by 12 percent in Table 1. "
                "We show that treated districts have lower dropout rates in Table 2. "
                "The main result is robust in the central specification. "
                "The effect is largest for poorer municipalities. "
                "Evidence suggests the program changed household investment."
            ),
            "introduction": "Table 1 reports baseline effects. Table 2 reports heterogeneity.",
        },
        "headline_table_selection": [
            {"item_id": "Table1", "title": "Main estimates"},
            {"item_id": "Table2", "title": "Heterogeneity"},
        ],
        "result_item_plans": [
            {"item_id": "Table1", "item_type": "table", "status": "completed"},
            {"item_id": "Table2", "item_type": "table", "status": "completed"},
        ],
        "comparisons": [
            {"metric_id": "t1_a", "table_name": "Table 1", "match": True},
            {"metric_id": "t1_b", "table_name": "Table 1", "match": False},
            {"metric_id": "t2_a", "table_name": "Table 2", "match": True},
        ],
        "manifest_total": 4,
        "compared_total": 3,
        "paper_visible_manifest_total": 4,
        "paper_visible_compared_total": 3,
        "coverage_pct": 75.0,
        "script_steps_completed": 2,
        "partial_results_available": True,
        "failure_records": [],
    }


def _alignment_payload():
    return {
        "findings": [
            {"status": "aligned", "severity": "info", "message": "Paper and code both reference fixed effects."},
            {"status": "mismatch", "severity": "medium", "message": "Paper discusses clustering, but code signals were weak."},
        ]
    }


def _robustness_payload():
    return {
        "checks": [
            {
                "name": "Window sensitivity",
                "summary": "Vary the main estimation window.",
                "category": "sample",
                "subcategory": "window_or_bandwidth",
                "status": "proposed",
            },
            {
                "name": "Inference sensitivity",
                "summary": "Use an alternate variance estimator.",
                "category": "inference",
                "subcategory": "variance_estimator",
                "status": "proposed",
            },
            {
                "name": "Controls sensitivity",
                "summary": "Toggle the main control bundle.",
                "category": "controls",
                "subcategory": "control_set",
                "status": "proposed",
            },
            {
                "name": "Sample sensitivity",
                "summary": "Use an alternate supported sample restriction.",
                "category": "sample",
                "subcategory": "subgroup_or_restriction",
                "status": "proposed",
            },
        ]
    }


def test_annotation_schema_creation_is_idempotent(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    RunCatalog(storage)
    RunCatalog(storage)

    with sqlite3.connect(storage.catalog_path) as connection:
        tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }

    assert "annotation_replication_papers" in tables
    assert "annotation_alignment_inconsistencies" in tables
    assert "annotation_robustness_checks" in tables
    assert "annotation_claims" in tables
    assert "annotation_claim_table_links" in tables


def test_annotation_schema_backfills_main_result_text_from_claim_details(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    RunCatalog(storage)
    with sqlite3.connect(storage.catalog_path) as connection:
        connection.execute(
            "INSERT INTO annotation_replication_papers (unique_id, paper_title) VALUES (?, ?)",
            ("10001", "Test paper"),
        )
        connection.execute(
            """
            INSERT INTO annotation_claims (
                unique_id, model, run_id, claim_rank, claim_text, claim_source
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            ("10001", 1, "run-test", 1, "The paper's main policy effect is positive.", "unit_test"),
        )

    RunCatalog(storage)
    with sqlite3.connect(storage.catalog_path) as connection:
        value = connection.execute(
            """
            SELECT main_result_1_m1_text
            FROM annotation_replication_papers
            WHERE unique_id = '10001'
            """
        ).fetchone()[0]

    assert value == "The paper's main policy effect is positive."


def test_build_important_claims_preserves_existing_main_results():
    results = _results("/tmp/paper.pdf")
    results["main_results"] = [
        {"claim_rank": 1, "claim_text": "Custom main result one.", "mapped_tables": ["Table1"]},
        {"claim_rank": 2, "claim_text": "Custom main result two.", "mapped_tables": ["Table2"]},
        {"claim_rank": 3, "claim_text": "Custom main result three.", "mapped_tables": ["Table1"]},
        {"claim_rank": 4, "claim_text": "Custom main result four.", "mapped_tables": ["Table2"]},
        {"claim_rank": 5, "claim_text": "Custom main result five.", "mapped_tables": ["Table1", "Table2"]},
    ]

    claims = build_important_claims(results)

    assert [claim["claim_text"] for claim in claims] == [
        "Custom main result one.",
        "Custom main result two.",
        "Custom main result three.",
        "Custom main result four.",
        "Custom main result five.",
    ]
    assert claims[4]["mapped_tables"] == ["Table1", "Table2"]


def test_build_important_claims_does_not_fallback_when_model_required():
    results = _results("/tmp/paper.pdf")
    results["important_claims"] = []
    results["important_claims_source"] = "model"
    results["claims_model_generated"] = False

    assert build_important_claims(results) == []


def test_record_annotation_outputs_populates_model_columns_and_claim_details(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    run_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )

    counts = catalog.record_annotation_outputs(
        run_context,
        _results(paper_path),
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    with sqlite3.connect(storage.catalog_path) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT exec_success_m1, exec_success_m2, identified_results_m1_n,
                   main_result_1_m1_text, main_result_5_m1_text, main_result_1_m2_text,
                   tables_m1_n, comparison_AIRE_t1_m1_n, match_AIRE_t1_m1_n,
                   incon_AIRE_m1_n, rob_AIRE_m1_n
            FROM annotation_replication_papers
            WHERE unique_id = '10001'
            """
        ).fetchone()
        claim_count = connection.execute(
            "SELECT COUNT(*) FROM annotation_claims WHERE unique_id = '10001' AND model = 1"
        ).fetchone()[0]
        max_links = connection.execute(
            """
            SELECT MAX(link_count)
            FROM (
                SELECT COUNT(*) AS link_count
                FROM annotation_claim_table_links
                WHERE unique_id = '10001' AND model = 1
                GROUP BY claim_rank
            )
            """
        ).fetchone()[0]

    assert counts["claims"] == 5
    assert row["exec_success_m1"] == 1
    assert row["exec_success_m2"] is None
    assert row["identified_results_m1_n"] == 5
    assert row["main_result_1_m1_text"].startswith("We find that the policy increases enrollment")
    assert row["main_result_5_m1_text"]
    assert row["main_result_1_m2_text"] is None
    assert row["tables_m1_n"] == 2
    assert row["comparison_AIRE_t1_m1_n"] == 2
    assert row["match_AIRE_t1_m1_n"] == 1
    assert row["incon_AIRE_m1_n"] == 1
    assert row["rob_AIRE_m1_n"] == 4
    assert claim_count == 5
    assert max_links <= 2


def test_annotation_execution_failure_ignores_recovered_attempts(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    run_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )
    results = _results(paper_path)
    results.update(
        {
            "completion_gate": "passed",
            "missing_total": 0,
            "status": "completed",
            "failure_records": [
                {
                    "severity": "missing_dependency",
                    "stage": "execution",
                    "tool": "run_planned_step",
                    "command": "step_01.do",
                    "likely_cause": "A package was unavailable on the first attempt.",
                    "recommended_fix": "Recovered by installing the package.",
                    "downstream_allowed": True,
                }
            ],
        }
    )

    catalog.record_annotation_outputs(
        run_context,
        results,
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    with sqlite3.connect(storage.catalog_path) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT exec_failure_type_m1, exec_failure_m1_note
            FROM annotation_replication_papers
            WHERE unique_id = '10001'
            """
        ).fetchone()

    assert row["exec_failure_type_m1"] == ""
    assert row["exec_failure_m1_note"] == ""


def test_annotation_execution_failure_note_keeps_exact_diagnosis(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    run_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )
    results = _results(paper_path, status="blocked")
    results.update(
        {
            "completion_gate": "blocked",
            "missing_total": 113,
            "compared_total": 0,
            "paper_visible_compared_total": 0,
            "coverage_pct": 0.0,
            "script_steps_completed": 0,
            "partial_results_available": False,
            "paper_items_blocked": 2,
            "unsupported_items": [
                {
                    "item_id": "Table2",
                    "title": "TABLE 2--FOREIGN SALES RATIO",
                    "evidence_kind": "unsupported_by_package",
                    "evidence_status": "blocked_unbound",
                    "unsupported_reason": "Selected item has no package-bound planned step or engine-verified current-run artifact.",
                }
            ],
            "failure_records": [
                {
                    "severity": "missing_dependency",
                    "stage": "execution",
                    "tool": "execute_code",
                    "command": "ad hoc schema probe",
                    "stderr_excerpt": "FileNotFoundError: [Errno 2] No such file or directory: 'stata-mp'",
                    "likely_cause": "A schema probe failed, but the final blocker is unsupported package evidence.",
                    "recommended_fix": "Provide executable analysis code for the selected table.",
                    "downstream_allowed": True,
                }
            ],
        }
    )

    catalog.record_annotation_outputs(
        run_context,
        results,
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    with sqlite3.connect(storage.catalog_path) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute(
            """
            SELECT exec_success_m1, exec_failure_type_m1, exec_failure_m1_note
            FROM annotation_replication_papers
            WHERE unique_id = '10001'
            """
        ).fetchone()

    assert row["exec_success_m1"] == 0
    assert row["exec_failure_type_m1"] == "8"
    assert "type=unsupported_by_package" in row["exec_failure_m1_note"]
    assert "where=evidence_binding/evidence_validator" in row["exec_failure_m1_note"]
    assert "no package-bound planned step" in row["exec_failure_m1_note"]
    assert "ad hoc schema probe" not in row["exec_failure_m1_note"]


def test_dual_model_annotation_updates_same_paper_row(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    gpt_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )
    opus_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="claude-opus-4-7",
        provider="anthropic",
    )

    catalog.record_annotation_outputs(
        gpt_context,
        _results(paper_path),
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )
    catalog.record_annotation_outputs(
        opus_context,
        _results(paper_path),
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    with sqlite3.connect(storage.catalog_path) as connection:
        row_count = connection.execute(
            "SELECT COUNT(*) FROM annotation_replication_papers"
        ).fetchone()[0]
        row = connection.execute(
            """
            SELECT exec_success_m1, exec_success_m2, run_id_m1, run_id_m2
            FROM annotation_replication_papers
            WHERE unique_id = '10001'
            """
        ).fetchone()

    assert row_count == 1
    assert row[0] == 1
    assert row[1] == 1
    assert row[2] == gpt_context.run_id
    assert row[3] == opus_context.run_id


def test_export_annotation_workbook_has_three_sheets_and_expected_headers(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    run_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )
    catalog.record_annotation_outputs(
        run_context,
        _results(paper_path),
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    workbook_path = catalog.export_annotation_workbook()
    workbook = load_workbook(workbook_path, read_only=True)

    assert workbook.sheetnames == [
        "database_1_replication",
        "database_2_alignment",
        "database_3_robustness",
    ]
    assert [cell.value for cell in next(workbook["database_1_replication"].iter_rows(max_row=1))] == REPLICATION_COLUMNS
    assert [cell.value for cell in next(workbook["database_2_alignment"].iter_rows(max_row=1))] == ALIGNMENT_COLUMNS
    assert [cell.value for cell in next(workbook["database_3_robustness"].iter_rows(max_row=1))] == ROBUSTNESS_COLUMNS
    first_replication_row = next(workbook["database_1_replication"].iter_rows(min_row=2, max_row=2, values_only=True))
    main_result_column = REPLICATION_COLUMNS.index("main_result_1_m1_text")
    assert first_replication_row[main_result_column].startswith("We find that the policy increases enrollment")
    workbook.close()


def test_export_annotation_workbook_sanitizes_control_characters(tmp_path):
    storage = StorageConfig(runs_root=str(tmp_path / "runs"))
    catalog = RunCatalog(storage)
    paper_path = _paper_path(tmp_path)
    run_context = catalog.create_run_context(
        paper_path=paper_path,
        model_name="gpt-5.5",
        provider="openai",
    )
    results = _results(paper_path, status="failed")
    results["failure_records"] = [
        {
            "severity": "missing_dependency",
            "stage": "execution",
            "tool": "run_planned_step",
            "stderr_excerpt": "\x1b[?1hstata output\x1b[0m with bad control \x07 char",
        }
    ]
    catalog.record_annotation_outputs(
        run_context,
        results,
        alignment_payload=_alignment_payload(),
        robustness_payload=_robustness_payload(),
    )

    workbook_path = catalog.export_annotation_workbook()
    workbook = load_workbook(workbook_path, read_only=True)
    row = next(workbook["database_1_replication"].iter_rows(min_row=2, max_row=2, values_only=True))
    diagnosis = row[REPLICATION_COLUMNS.index("failure_diagnosis_m1")]

    assert "\x1b" not in diagnosis
    assert "\x07" not in diagnosis
    assert "stata output" in diagnosis
    workbook.close()


def test_robustness_rows_block_when_model_generated_checks_are_missing():
    rows = build_robustness_rows(
        {"paper_id": "10001", "paper_metadata": {"title": "Test paper"}},
        {"checks": []},
        model_index=1,
    )

    assert len(rows) == 4
    assert {row["status"] for row in rows} == {"blocked"}
    assert {row["rob_AIRE_subcat"] for row in rows} == {"robustness_agent_output_missing"}


def test_robustness_rows_record_blocked_failure_without_model_generated_checks():
    rows = build_robustness_rows(
        {
            "paper_id": "10001",
            "paper_metadata": {"title": "Test paper"},
            "status": "blocked",
            "completion_gate": "blocked",
            "missing_total": 10,
            "paper_items_blocked": 1,
            "unsupported_items": [
                {
                    "item_id": "Table2",
                    "evidence_kind": "unsupported_by_package",
                    "evidence_status": "blocked_unbound",
                    "unsupported_reason": "Selected item has no package-bound planned step or engine-verified current-run artifact.",
                }
            ],
        },
        {"checks": []},
        model_index=1,
    )

    assert len(rows) == 4
    assert {row["status"] for row in rows} == {"blocked"}
    assert {row["rob_AIRE_cat"] for row in rows} == {"blocked"}
    assert "type=unsupported_by_package" in rows[0]["rob_AIRE_des"]
